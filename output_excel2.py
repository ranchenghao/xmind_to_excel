import xmindparser
import openpyxl


def read_xmind():
    xmindparser.config = {
        'showTopicId': True,  # 是否展示主题ID
        'hideEmptyValue': True  # 是否隐藏空值
    }

    filePath = '调班单调整.xmind'

    # 解析成json数据类型
    content = xmindparser.xmind_to_dict(filePath)
    return content


def get_value_from_json(tdict, tem_list, case_data):
    """
    从Json中获取key值，
    :param key:
    :param tdict:
    :param tem_list:
    :return:
    """
    tem_list.append(tdict)
    if not isinstance(tdict, dict):
        return tdict + "is not dict"
    elif tdict.get('topics') is None:
        if len(tem_list[-3]) == 3:
            print(tem_list)
            case_data.append(tem_list[-3])
        else:
            print("用例编写不规范")
            print(tem_list[-1])
        # yield tem_list[-3]
        # tem_list = []
    else:
        for value in tdict.values():
            if isinstance(value, dict):
                get_value_from_json(value, tem_list, case_data)
            elif isinstance(value, (list, tuple)):
                _get_value(value, tem_list, case_data)
    return case_data


def _get_value(tdict, tem_list, case_data):
    """

    :param key:
    :param tdict:
    :param tem_list:
    :return:
    """
    for value in tdict:
        if isinstance(value, (list, tuple)):
            _get_value(tdict)
        elif isinstance(value, dict):
            get_value_from_json(value, tem_list, case_data)


def write_excel(data):
    wb = openpyxl.load_workbook('测试用例模板.xlsx')
    ws = wb['模板']
    r = 2
    for case in get_value_from_json(data[0]['topic'], [], []):
        ws['B{}'.format(r)] = case['title']
        if case.get('topics'):
            ws['E{}'.format(r)] = case.get('topics')[0]['title']
        if case['topics'][0].get('topics'):
            ws['F{}'.format(r)] = case['topics'][0]['topics'][0]['title']
        if case['topics'][0]['topics'][0].get('makers'):
            ws['I{}'.format(r)] = '高'
        r += 1
    wb.save(u'output.xlsx')
    wb.close()


data = read_xmind()
write_excel(data)
